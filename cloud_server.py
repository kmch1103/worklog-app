from flask import Flask, request, jsonify, render_template, send_file
import sqlite3
import json
import os
import tempfile
import uuid
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference

app = Flask(__name__)
DB = "worklog.db"


def db():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    return conn


def rows_to_dicts(rows):
    return [dict(r) for r in rows]


def ensure_column(cur, table_name, column_name, column_def):
    cols = [r["name"] for r in cur.execute(f"PRAGMA table_info({table_name})").fetchall()]
    if column_name not in cols:
        cur.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_def}")


def normalize_name(value):
    return (value or "").strip()


def sync_material_option(cur, material_name):
    name = normalize_name(material_name)
    if not name:
        return

    row = cur.execute(
        "SELECT id FROM options_materials WHERE name = ?",
        (name,)
    ).fetchone()

    if not row:
        cur.execute(
            "INSERT INTO options_materials (name) VALUES (?)",
            (name,)
        )


def safe_float(value, default=0):
    try:
        if value is None or value == "":
            return default
        return float(value)
    except Exception:
        return default


def row_value(row, key, default=""):
    try:
        value = row[key]
        return default if value is None else value
    except Exception:
        return default


def parse_old_materials(raw_text, material_price_map):
    raw = (raw_text or "").strip()
    if not raw:
        return []

    parts = [p.strip() for p in raw.split(";") if p.strip()]
    result = []

    for part in parts:
        tokens = [t.strip() for t in part.split("|")]
        name = tokens[0] if len(tokens) >= 1 else ""
        qty = safe_float(tokens[1], 0) if len(tokens) >= 2 else 0
        unit = tokens[2] if len(tokens) >= 3 else ""
        price = safe_float(material_price_map.get(name), 0)

        if not name:
            continue

        result.append({
            "id": "",
            "name": name,
            "unit": unit,
            "price": price,
            "qty": qty,
            "method": ""
        })

    return result


def parse_old_labor_rows(row):
    raw = str(row_value(row, "인력내역", "") or "").strip()
    rows = []

    if raw:
        for chunk in raw.split(";"):
            chunk = chunk.strip()
            if not chunk:
                continue

            tokens = [t.strip() for t in chunk.split("|")]
            labor_type = tokens[0] if len(tokens) >= 1 else ""
            count = int(safe_float(tokens[1], 0)) if len(tokens) >= 2 else 0
            price = safe_float(tokens[2], 0) if len(tokens) >= 3 else 0
            note = tokens[3] if len(tokens) >= 4 else ""
            amount = count * price

            if count > 0 or price > 0 or note:
                rows.append({
                    "type": labor_type,
                    "count": count,
                    "price": price,
                    "amount": amount,
                    "method": "",
                    "note": note
                })

    if rows:
        return rows

    legacy_map = [
        ("남자", "남자수", "남자단가"),
        ("여자", "여자수", "여자단가"),
        ("기타", "기타수", "기타단가")
    ]

    for labor_type, count_key, price_key in legacy_map:
        count = int(safe_float(row_value(row, count_key, 0), 0))
        price = safe_float(row_value(row, price_key, 0), 0)
        amount = count * price
        if count > 0 or price > 0:
            rows.append({
                "type": labor_type,
                "count": count,
                "price": price,
                "amount": amount,
                "method": "",
                "note": ""
            })

    return rows


def build_import_money(materials, labor_rows, legacy_labor_cost):
    material_total = sum(safe_float(item.get("qty"), 0) * safe_float(item.get("price"), 0) for item in materials)
    labor_total = sum(safe_float(item.get("amount"), 0) for item in labor_rows)

    if labor_total <= 0:
        labor_total = safe_float(legacy_labor_cost, 0)

    total_amount = labor_total + material_total
    if total_amount <= 0:
        return None

    if labor_total > 0 and material_total > 0:
        money_type = "인건비+자재비"
    elif labor_total > 0:
        money_type = "인건비"
    elif material_total > 0:
        money_type = "자재비"
    else:
        money_type = "기타"

    return {
        "type": money_type,
        "total_amount": total_amount,
        "labor_total": labor_total,
        "material_total": material_total,
        "other_total": 0,
        "method": "",
        "note": "기존 DB 가져오기"
    }


def clear_current_data(cur):
    cur.execute("DELETE FROM works")
    cur.execute("DELETE FROM materials")
    for t in ["weather", "crops", "tasks", "pests", "materials", "machines"]:
        cur.execute(f"DELETE FROM options_{t}")


def import_old_options(old_conn, cur, old_table, new_table, old_column):
    try:
        rows = old_conn.execute(f"SELECT {old_column} FROM '{old_table}'").fetchall()
        for row in rows:
            name = normalize_name(row[0])
            if name:
                cur.execute(f"INSERT OR IGNORE INTO {new_table} (name) VALUES (?)", (name,))
    except Exception:
        pass


def import_old_db_into_current(uploaded_db_path):
    old_conn = sqlite3.connect(uploaded_db_path)
    old_conn.row_factory = sqlite3.Row

    conn = db()
    cur = conn.cursor()

    try:
        material_rows = old_conn.execute("SELECT 자재명, 단위, 가격, 재고 FROM 자재").fetchall()
        material_price_map = {}
        for row in material_rows:
            name = normalize_name(row_value(row, "자재명", ""))
            if name:
                material_price_map[name] = safe_float(row_value(row, "가격", 0), 0)

        clear_current_data(cur)

        for row in material_rows:
            name = normalize_name(row_value(row, "자재명", ""))
            unit = normalize_name(row_value(row, "단위", ""))
            unit_price = safe_float(row_value(row, "가격", 0), 0)
            stock_qty = safe_float(row_value(row, "재고", 0), 0)

            if not name:
                continue

            cur.execute("""
                INSERT INTO materials (name, unit, stock_qty, unit_price, price_last_year, price_this_year, memo)
                VALUES (?, ?, ?, ?, 0, 0, '')
            """, (name, unit, stock_qty, unit_price))
            sync_material_option(cur, name)

        import_old_options(old_conn, cur, "옵션_날씨", "options_weather", "항목")
        import_old_options(old_conn, cur, "옵션_작물", "options_crops", "항목")
        import_old_options(old_conn, cur, "옵션_작업내용", "options_tasks", "항목")
        import_old_options(old_conn, cur, "옵션_기계", "options_machines", "항목")
        import_old_options(old_conn, cur, "병충해", "options_pests", "이름")

        work_rows = old_conn.execute("SELECT * FROM 작업일지 ORDER BY 날짜 ASC, 번호 ASC").fetchall()

        for row in work_rows:
            materials = parse_old_materials(row_value(row, "사용자재", ""), material_price_map)
            labor_rows = parse_old_labor_rows(row)
            money = build_import_money(materials, labor_rows, row_value(row, "인건비", 0))

            memo_obj = {
                "memo_text": str(row_value(row, "비고", "") or ""),
                "repeat_days": 1,
                "start_time": str(row_value(row, "시작시간", "") or ""),
                "end_time": str(row_value(row, "종료시간", "") or ""),
                "materials": materials,
                "labor_rows": labor_rows,
                "work_hours": safe_float(row_value(row, "작업시간", 0), 0),
                "money": money,
                "legacy": {
                    "created_at": str(row_value(row, "생성시각", "") or ""),
                    "updated_at": str(row_value(row, "수정시각", "") or ""),
                    "work_list": str(row_value(row, "작업목록", "") or "")
                }
            }

            cur.execute("""
                INSERT INTO works (
                    start_date, end_date, weather, task_name,
                    crops, pests, machines, work_hours, memo
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                str(row_value(row, "날짜", "") or ""),
                str(row_value(row, "종료날짜", "") or row_value(row, "날짜", "") or ""),
                str(row_value(row, "날씨", "") or ""),
                str(row_value(row, "작업내용", "") or ""),
                str(row_value(row, "작물", "") or ""),
                str(row_value(row, "적용병충해", "") or ""),
                str(row_value(row, "사용기계", "") or ""),
                safe_float(row_value(row, "작업시간", 0), 0),
                json.dumps(memo_obj, ensure_ascii=False)
            ))

        conn.commit()
        return {
            "ok": True,
            "imported": {
                "works": len(work_rows),
                "materials": len(material_rows)
            }
        }
    except Exception as e:
        conn.rollback()
        return {"ok": False, "error": str(e)}
    finally:
        conn.close()
        old_conn.close()





def parse_memo_json(raw_value):
    if isinstance(raw_value, dict):
        return raw_value
    raw_text = (raw_value or '').strip()
    if not raw_text:
        return {}
    try:
        parsed = json.loads(raw_text)
        return parsed if isinstance(parsed, dict) else {}
    except Exception:
        return {}


def apply_material_stock_change(conn, materials, mode='apply'):
    cur = conn.cursor()

    for item in materials or []:
        name = normalize_name(item.get('name'))
        qty = safe_float(item.get('qty'), 0)
        material_type = normalize_name(item.get('material_type') or item.get('kind') or '재고형')
        action = normalize_name(item.get('action') or item.get('behavior') or item.get('type') or '사용')

        if material_type == '비용형':
            continue

        if not name or qty == 0:
            continue

        stock_delta = 0
        if action == '구입':
            stock_delta = qty
        elif action == '사용':
            stock_delta = -qty
        elif action == '반품':
            stock_delta = -qty
        else:
            continue

        if mode == 'revert':
            stock_delta = -stock_delta

        cur.execute(
            'UPDATE materials SET stock_qty = COALESCE(stock_qty, 0) + ? WHERE name = ?',
            (stock_delta, name)
        )

def current_timestamp():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def get_season_row(conn, season_id):
    if season_id in (None, '', 'all'):
        return None

    if season_id == 'current':
        return conn.execute("SELECT * FROM seasons WHERE is_current = 1 ORDER BY id DESC LIMIT 1").fetchone()

    try:
        sid = int(season_id)
    except Exception:
        return None

    return conn.execute("SELECT * FROM seasons WHERE id = ?", (sid,)).fetchone()


def apply_season_filter_sql(base_sql, season_row, date_column='start_date'):
    if not season_row:
        return base_sql, ()
    return f"{base_sql} WHERE {date_column} BETWEEN ? AND ?", (season_row['start_date'], season_row['end_date'])


def season_payload(row):
    if not row:
        return None
    return {
        'id': row['id'],
        'season_name': row['season_name'],
        'start_date': row['start_date'],
        'end_date': row['end_date'],
        'is_current': row['is_current'],
        'note': row['note'],
        'created_at': row['created_at'],
    }

def init_db():
    conn = db()
    cur = conn.cursor()

    # works
    cur.execute("""
    CREATE TABLE IF NOT EXISTS works (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        start_date TEXT,
        end_date TEXT,
        weather TEXT,
        task_name TEXT,
        crops TEXT,
        pests TEXT,
        machines TEXT,
        work_hours REAL DEFAULT 0,
        memo TEXT DEFAULT ''
    )
    """)

    ensure_column(cur, "works", "end_date", "TEXT")
    ensure_column(cur, "works", "weather", "TEXT")
    ensure_column(cur, "works", "task_name", "TEXT")
    ensure_column(cur, "works", "crops", "TEXT")
    ensure_column(cur, "works", "pests", "TEXT")
    ensure_column(cur, "works", "machines", "TEXT")
    ensure_column(cur, "works", "work_hours", "REAL DEFAULT 0")
    ensure_column(cur, "works", "memo", "TEXT DEFAULT ''")
    ensure_column(cur, "works", "task_category", "TEXT DEFAULT ''")


    
    # plans
    cur.execute("""
    CREATE TABLE IF NOT EXISTS plans (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        plan_date TEXT,
        title TEXT,
        details TEXT,
        status TEXT DEFAULT 'planned'
    )
    """)

    # materials
    cur.execute("""
    CREATE TABLE IF NOT EXISTS materials (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        unit TEXT DEFAULT '',
        stock_qty REAL DEFAULT 0,
        unit_price REAL DEFAULT 0,
        price_last_year REAL DEFAULT 0,
        price_this_year REAL DEFAULT 0,
        memo TEXT DEFAULT ''
    )
    """)

    ensure_column(cur, "materials", "unit", "TEXT DEFAULT ''")
    ensure_column(cur, "materials", "stock_qty", "REAL DEFAULT 0")
    ensure_column(cur, "materials", "unit_price", "REAL DEFAULT 0")
    ensure_column(cur, "materials", "price_last_year", "REAL DEFAULT 0")
    ensure_column(cur, "materials", "price_this_year", "REAL DEFAULT 0")
    ensure_column(cur, "materials", "memo", "TEXT DEFAULT ''")

    # option tables
    for t in ["weather", "crops", "tasks", "pests", "materials", "machines"]:
        cur.execute(f"""
        CREATE TABLE IF NOT EXISTS options_{t} (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE
        )
        """)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS options_task_categories (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL UNIQUE
    )
    """)
    ensure_column(cur, "options_tasks", "category_name", "TEXT DEFAULT ''")
    ensure_column(cur, "options_pests", "recommended_materials", "TEXT DEFAULT ''")

    # seasons
    cur.execute("""
    CREATE TABLE IF NOT EXISTS seasons (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        season_name TEXT NOT NULL,
        start_date TEXT NOT NULL,
        end_date TEXT NOT NULL,
        is_current INTEGER DEFAULT 0,
        note TEXT DEFAULT '',
        created_at TEXT DEFAULT ''
    )
    """)
    ensure_column(cur, "seasons", "note", "TEXT DEFAULT ''")
    ensure_column(cur, "seasons", "created_at", "TEXT DEFAULT ''")

    # incomes
    cur.execute("""
    CREATE TABLE IF NOT EXISTS incomes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        income_date TEXT,
        income_type TEXT DEFAULT '',
        amount REAL DEFAULT 0,
        method TEXT DEFAULT '',
        note TEXT DEFAULT '',
        created_at TEXT DEFAULT ''
    )
    """)
    ensure_column(cur, "incomes", "income_date", "TEXT")
    ensure_column(cur, "incomes", "income_type", "TEXT DEFAULT ''")
    ensure_column(cur, "incomes", "amount", "REAL DEFAULT 0")
    ensure_column(cur, "incomes", "method", "TEXT DEFAULT ''")
    ensure_column(cur, "incomes", "note", "TEXT DEFAULT ''")
    ensure_column(cur, "incomes", "created_at", "TEXT DEFAULT ''")

    # favorite work templates
    cur.execute("""
    CREATE TABLE IF NOT EXISTS favorite_work_templates (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL UNIQUE,
        template_json TEXT DEFAULT '{}',
        created_at TEXT DEFAULT '',
        updated_at TEXT DEFAULT ''
    )
    """)
    ensure_column(cur, "favorite_work_templates", "template_json", "TEXT DEFAULT '{}' ")
    ensure_column(cur, "favorite_work_templates", "created_at", "TEXT DEFAULT ''")
    ensure_column(cur, "favorite_work_templates", "updated_at", "TEXT DEFAULT ''")

    conn.commit()
    conn.close()


init_db()


@app.route("/")
def index():
    return render_template("index.html")


# =========================
# WORKS
# =========================
@app.route("/api/works", methods=["GET"])
def get_works():
    conn = db()
    season_id = (request.args.get("season_id") or "").strip()
    season = get_season_row(conn, season_id)
    sql, params = apply_season_filter_sql("SELECT * FROM works", season, "start_date")
    rows = conn.execute(sql + " ORDER BY start_date DESC, id DESC", params).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/works", methods=["POST"])

def create_work():
    data = request.get_json(force=True) or {}

    start_date = (data.get("start_date") or "").strip()
    end_date = (data.get("end_date") or start_date).strip()
    task_name = (data.get("task_name") or "").strip()

    if not start_date:
        return jsonify({"ok": False, "error": "시작일이 없습니다."}), 400
    if not task_name:
        return jsonify({"ok": False, "error": "세부작업이 없습니다."}), 400

    conn = db()
    try:
        memo_text = data.get("memo", "")
        memo_obj = parse_memo_json(memo_text)
        apply_material_stock_change(conn, memo_obj.get("materials", []), mode='apply')

        cur = conn.cursor()
        cur.execute("""
            INSERT INTO works (
                start_date, end_date, weather, task_category, task_name,
                crops, pests, machines, work_hours, memo
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            start_date,
            end_date,
            data.get("weather", ""),
            data.get("task_category", ""),
            task_name,
            data.get("crops", ""),
            data.get("pests", ""),
            data.get("machines", ""),
            float(data.get("work_hours") or 0),
            memo_text
        ))
        conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"ok": False, "error": f"작업 저장 오류: {e}"}), 500
    finally:
        conn.close()



@app.route("/api/works/<int:work_id>", methods=["PUT"])

def update_work(work_id):
    data = request.get_json(force=True) or {}

    start_date = (data.get("start_date") or "").strip()
    end_date = (data.get("end_date") or start_date).strip()
    task_name = (data.get("task_name") or "").strip()

    if not start_date:
        return jsonify({"ok": False, "error": "시작일이 없습니다."}), 400
    if not task_name:
        return jsonify({"ok": False, "error": "세부작업이 없습니다."}), 400

    conn = db()
    try:
        old_row = conn.execute("SELECT memo FROM works WHERE id = ?", (work_id,)).fetchone()
        if old_row:
            old_memo = parse_memo_json(old_row["memo"])
            apply_material_stock_change(conn, old_memo.get("materials", []), mode='revert')

        memo_text = data.get("memo", "")
        memo_obj = parse_memo_json(memo_text)
        apply_material_stock_change(conn, memo_obj.get("materials", []), mode='apply')

        cur = conn.cursor()
        cur.execute("""
            UPDATE works
            SET start_date = ?,
                end_date = ?,
                weather = ?,
                task_category = ?,
                task_name = ?,
                crops = ?,
                pests = ?,
                machines = ?,
                work_hours = ?,
                memo = ?
            WHERE id = ?
        """, (
            start_date,
            end_date,
            data.get("weather", ""),
            data.get("task_category", ""),
            task_name,
            data.get("crops", ""),
            data.get("pests", ""),
            data.get("machines", ""),
            float(data.get("work_hours") or 0),
            memo_text,
            work_id
        ))
        conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"ok": False, "error": f"작업 수정 오류: {e}"}), 500
    finally:
        conn.close()



@app.route("/api/works/<int:work_id>", methods=["DELETE"])
def delete_work(work_id):
    conn = db()
    try:
        old_row = conn.execute("SELECT memo FROM works WHERE id = ?", (work_id,)).fetchone()
        if old_row:
            old_memo = parse_memo_json(old_row["memo"])
            apply_material_stock_change(conn, old_memo.get("materials", []), mode='revert')

        conn.execute("DELETE FROM works WHERE id = ?", (work_id,))
        conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"ok": False, "error": f"작업 삭제 오류: {e}"}), 500
    finally:
        conn.close()


# =========================
# PLANS
# =========================
@app.route("/api/plans", methods=["GET"])
def get_plans():
    conn = db()
    rows = conn.execute(
        "SELECT * FROM plans ORDER BY plan_date DESC, id DESC"
    ).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/plans", methods=["POST"])
def create_plan():
    data = request.get_json(force=True) or {}

    conn = db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO plans (plan_date, title, details, status)
        VALUES (?, ?, ?, ?)
    """, (
        data.get("plan_date", ""),
        data.get("title", ""),
        data.get("details", ""),
        data.get("status", "planned")
    ))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/plans/<int:plan_id>", methods=["PUT"])
def update_plan(plan_id):
    data = request.get_json(force=True) or {}

    conn = db()
    cur = conn.cursor()
    cur.execute("""
        UPDATE plans
        SET plan_date = ?,
            title = ?,
            details = ?,
            status = ?
        WHERE id = ?
    """, (
        data.get("plan_date", ""),
        data.get("title", ""),
        data.get("details", ""),
        data.get("status", "planned"),
        plan_id
    ))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/plans/<int:plan_id>", methods=["DELETE"])
def delete_plan(plan_id):
    conn = db()
    conn.execute("DELETE FROM plans WHERE id = ?", (plan_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# =========================
# OPTIONS
# =========================
@app.route("/api/options", methods=["GET"])
def get_options():
    conn = db()

    # 프론트에서 지금 쓰는 것만 내려줌
    result = {}
    for t in ["weather", "crops", "machines", "task_categories"]:
        table = f"options_{t}"
        rows = conn.execute(
            f"SELECT id, name FROM {table} ORDER BY name"
        ).fetchall()
        result[t] = rows_to_dicts(rows)

    task_rows = conn.execute("""
        SELECT id, name, COALESCE(category_name, '') AS category_name
        FROM options_tasks
        ORDER BY category_name, name
    """).fetchall()
    result["tasks"] = rows_to_dicts(task_rows)

    pest_rows = conn.execute("""
        SELECT id, name, COALESCE(recommended_materials, '') AS recommended_materials
        FROM options_pests
        ORDER BY name
    """).fetchall()
    result["pests"] = rows_to_dicts(pest_rows)

    conn.close()
    return jsonify(result)


@app.route("/api/options/<option_type>", methods=["POST"])
def create_option(option_type):
    if option_type not in ["weather", "crops", "task_categories", "tasks", "pests", "machines", "materials"]:
        return jsonify({"ok": False, "error": "invalid option type"}), 400

    data = request.get_json(force=True) or {}
    name = normalize_name(data.get("name"))
    recommended_materials = normalize_name(data.get("recommended_materials"))
    category_name = normalize_name(data.get("category_name"))
    category_name = normalize_name(data.get("category_name"))
    if not name:
        return jsonify({"ok": False, "error": "name required"}), 400

    conn = db()
    cur = conn.cursor()

    if option_type == "tasks":
        cur.execute(
            "INSERT OR IGNORE INTO options_tasks (name, category_name) VALUES (?, ?)",
            (name, category_name)
        )
        cur.execute(
            "UPDATE options_tasks SET category_name = ? WHERE name = ?",
            (category_name, name)
        )
    elif option_type == "pests":
        cur.execute(
            "INSERT OR IGNORE INTO options_pests (name, recommended_materials) VALUES (?, ?)",
            (name, recommended_materials)
        )
        cur.execute(
            "UPDATE options_pests SET recommended_materials = ? WHERE name = ?",
            (recommended_materials, name)
        )
    else:
        cur.execute(
            f"INSERT OR IGNORE INTO options_{option_type} (name) VALUES (?)",
            (name,)
        )

    # materials 옵션을 직접 만들면 materials 테이블에도 최소 등록
    if option_type == "materials":
        row = cur.execute(
            "SELECT id FROM materials WHERE name = ?",
            (name,)
        ).fetchone()
        if not row:
            cur.execute("""
                INSERT INTO materials (name, unit, stock_qty, unit_price, memo)
                VALUES (?, '', 0, 0, '')
            """, (name,))

    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/options/<option_type>/<int:option_id>", methods=["PUT"])

def update_option(option_type, option_id):
    if option_type not in ["weather", "crops", "task_categories", "tasks", "pests", "machines", "materials"]:
        return jsonify({"ok": False, "error": "invalid option type"}), 400

    data = request.get_json(force=True) or {}
    new_name = normalize_name(data.get("name"))
    recommended_materials = normalize_name(data.get("recommended_materials"))
    category_name = normalize_name(data.get("category_name"))
    if not new_name:
        return jsonify({"ok": False, "error": "name required"}), 400

    conn = db()
    cur = conn.cursor()

    table = "options_task_categories" if option_type == "task_categories" else f"options_{option_type}"
    old_row = cur.execute(
        f"SELECT name FROM {table} WHERE id = ?",
        (option_id,)
    ).fetchone()

    if option_type == "tasks":
        cur.execute(
            "UPDATE options_tasks SET name = ?, category_name = ? WHERE id = ?",
            (new_name, category_name, option_id)
        )
    elif option_type == "pests":
        cur.execute(
            "UPDATE options_pests SET name = ?, recommended_materials = ? WHERE id = ?",
            (new_name, recommended_materials, option_id)
        )
    else:
        cur.execute(
            f"UPDATE {table} SET name = ? WHERE id = ?",
            (new_name, option_id)
        )

    if option_type == "task_categories" and old_row:
        old_name = old_row["name"]
        cur.execute(
            "UPDATE options_tasks SET category_name = ? WHERE category_name = ?",
            (new_name, old_name)
        )

    if option_type == "materials" and old_row:
        old_name = old_row["name"]
        cur.execute(
            "UPDATE materials SET name = ? WHERE name = ?",
            (new_name, old_name)
        )

    conn.commit()
    conn.close()
    return jsonify({"ok": True})



@app.route("/api/options/<option_type>/<int:option_id>", methods=["DELETE"])

def delete_option(option_type, option_id):
    if option_type not in ["weather", "crops", "task_categories", "tasks", "pests", "machines", "materials"]:
        return jsonify({"ok": False, "error": "invalid option type"}), 400

    conn = db()
    cur = conn.cursor()

    table = "options_task_categories" if option_type == "task_categories" else f"options_{option_type}"
    row = cur.execute(
        f"SELECT name FROM {table} WHERE id = ?",
        (option_id,)
    ).fetchone()

    cur.execute(
        f"DELETE FROM {table} WHERE id = ?",
        (option_id,)
    )

    if option_type == "task_categories" and row:
        cur.execute("UPDATE options_tasks SET category_name = '' WHERE category_name = ?", (row["name"],))

    if option_type == "materials" and row:
        cur.execute(
            "DELETE FROM materials WHERE name = ?",
            (row["name"],)
        )

    conn.commit()
    conn.close()
    return jsonify({"ok": True})



# =========================
# MATERIALS
# =========================
@app.route("/api/materials", methods=["GET"])
def get_materials():
    conn = db()
    rows = conn.execute("""
        SELECT id, name, unit, stock_qty, unit_price, price_last_year, price_this_year, memo
        FROM materials
        ORDER BY
          CASE WHEN COALESCE(stock_qty, 0) > 0 THEN 0 ELSE 1 END,
          name
    """).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/materials", methods=["POST"])
def create_material():
    data = request.get_json(force=True) or {}

    name = normalize_name(data.get("name"))
    unit = normalize_name(data.get("unit"))
    stock_qty = float(data.get("stock_qty") or 0)
    unit_price = float(data.get("unit_price") or 0)
    price_last_year = float(data.get("price_last_year") or 0)
    price_this_year = float(data.get("price_this_year") or 0)
    memo = normalize_name(data.get("memo"))

    if unit_price <= 0 and price_this_year > 0:
        unit_price = price_this_year

    if not name:
        return jsonify({"ok": False, "error": "name required"}), 400

    conn = db()
    cur = conn.cursor()

    row = cur.execute(
        "SELECT id FROM materials WHERE name = ?",
        (name,)
    ).fetchone()

    if row:
        cur.execute("""
            UPDATE materials
            SET unit = ?, stock_qty = ?, unit_price = ?, price_last_year = ?, price_this_year = ?, memo = ?
            WHERE id = ?
        """, (unit, stock_qty, unit_price, price_last_year, price_this_year, memo, row["id"]))
    else:
        cur.execute("""
            INSERT INTO materials (name, unit, stock_qty, unit_price, price_last_year, price_this_year, memo)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (name, unit, stock_qty, unit_price, price_last_year, price_this_year, memo))

    sync_material_option(cur, name)

    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/materials/<int:material_id>", methods=["PUT"])
def update_material(material_id):
    data = request.get_json(force=True) or {}

    name = normalize_name(data.get("name"))
    unit = normalize_name(data.get("unit"))
    stock_qty = float(data.get("stock_qty") or 0)
    unit_price = float(data.get("unit_price") or 0)
    price_last_year = float(data.get("price_last_year") or 0)
    price_this_year = float(data.get("price_this_year") or 0)
    memo = normalize_name(data.get("memo"))

    if unit_price <= 0 and price_this_year > 0:
        unit_price = price_this_year

    if not name:
      return jsonify({"ok": False, "error": "name required"}), 400

    conn = db()
    cur = conn.cursor()

    old_row = cur.execute(
        "SELECT name FROM materials WHERE id = ?",
        (material_id,)
    ).fetchone()

    cur.execute("""
        UPDATE materials
        SET name = ?, unit = ?, stock_qty = ?, unit_price = ?, price_last_year = ?, price_this_year = ?, memo = ?
        WHERE id = ?
    """, (name, unit, stock_qty, unit_price, price_last_year, price_this_year, memo, material_id))

    if old_row:
        old_name = old_row["name"]
        opt_row = cur.execute(
            "SELECT id FROM options_materials WHERE name = ?",
            (old_name,)
        ).fetchone()
        if opt_row:
            cur.execute(
                "UPDATE options_materials SET name = ? WHERE id = ?",
                (name, opt_row["id"])
            )
        else:
            sync_material_option(cur, name)
    else:
        sync_material_option(cur, name)

    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/materials/<int:material_id>", methods=["DELETE"])
def delete_material(material_id):
    conn = db()
    cur = conn.cursor()

    row = cur.execute(
        "SELECT name FROM materials WHERE id = ?",
        (material_id,)
    ).fetchone()

    cur.execute("DELETE FROM materials WHERE id = ?", (material_id,))

    if row:
        cur.execute(
            "DELETE FROM options_materials WHERE name = ?",
            (row["name"],)
        )

    conn.commit()
    conn.close()
    return jsonify({"ok": True})





# =========================
# SEASONS
# =========================
@app.route("/api/seasons", methods=["GET"])
def get_seasons():
    conn = db()
    rows = conn.execute("SELECT * FROM seasons ORDER BY start_date DESC, id DESC").fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/seasons", methods=["POST"])
def create_season():
    data = request.get_json(force=True) or {}
    season_name = normalize_name(data.get("season_name") or data.get("name"))
    start_date = normalize_name(data.get("start_date"))
    end_date = normalize_name(data.get("end_date"))
    note = (data.get("note") or "").strip()
    is_current_raw = data.get("is_current")
    is_current = 1 if str(is_current_raw).lower() in ["1", "true", "yes", "y", "on"] else 0

    if not season_name or not start_date or not end_date:
        return jsonify({"ok": False, "error": "season_name, start_date, end_date required"}), 400

    conn = db()
    cur = conn.cursor()
    if is_current:
        cur.execute("UPDATE seasons SET is_current = 0")
    cur.execute("""
        INSERT INTO seasons (season_name, start_date, end_date, is_current, note, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (season_name, start_date, end_date, is_current, note, current_timestamp()))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/seasons/<int:season_id>", methods=["PUT"])
def update_season(season_id):
    data = request.get_json(force=True) or {}
    season_name = normalize_name(data.get("season_name") or data.get("name"))
    start_date = normalize_name(data.get("start_date"))
    end_date = normalize_name(data.get("end_date"))
    note = (data.get("note") or "").strip()
    is_current_raw = data.get("is_current")
    is_current = 1 if str(is_current_raw).lower() in ["1", "true", "yes", "y", "on"] else 0

    if not season_name or not start_date or not end_date:
        return jsonify({"ok": False, "error": "season_name, start_date, end_date required"}), 400

    conn = db()
    cur = conn.cursor()
    if is_current:
        cur.execute("UPDATE seasons SET is_current = 0")
    cur.execute("""
        UPDATE seasons
        SET season_name = ?, start_date = ?, end_date = ?, is_current = ?, note = ?
        WHERE id = ?
    """, (season_name, start_date, end_date, is_current, note, season_id))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/seasons/<int:season_id>", methods=["DELETE"])
def delete_season(season_id):
    conn = db()
    conn.execute("DELETE FROM seasons WHERE id = ?", (season_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/seasons/<int:season_id>/set_current", methods=["PUT"])
def set_current_season(season_id):
    conn = db()
    cur = conn.cursor()
    cur.execute("UPDATE seasons SET is_current = 0")
    cur.execute("UPDATE seasons SET is_current = 1 WHERE id = ?", (season_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/seasons/current", methods=["GET"])
def get_current_season():
    conn = db()
    row = get_season_row(conn, 'current')
    conn.close()
    return jsonify(season_payload(row) or {})


@app.route("/api/seasons/<int:season_id>/backup", methods=["GET"])
def backup_season(season_id):
    conn = db()
    season = get_season_row(conn, str(season_id))
    if not season:
        conn.close()
        return jsonify({"ok": False, "error": "season not found"}), 404

    works_sql, works_params = apply_season_filter_sql("SELECT * FROM works", season, "start_date")
    work_rows = conn.execute(works_sql + " ORDER BY start_date ASC, id ASC", works_params).fetchall()

    money_rows = []
    for row in work_rows:
        item = dict(row)
        try:
            memo = json.loads(item.get("memo") or "{}")
        except Exception:
            memo = {}
        money = memo.get("money") or {}
        if money:
            money_rows.append({
                "date": item.get("start_date", ""),
                "task_name": item.get("task_name", ""),
                "type": money.get("type", ""),
                "total": float(money.get("total_amount") or money.get("amount") or 0),
                "method": money.get("method", ""),
                "note": money.get("note", "")
            })

    materials = rows_to_dicts(conn.execute("SELECT * FROM materials ORDER BY name").fetchall())
    options = {}
    for t in ["weather", "crops", "tasks", "machines"]:
        options[t] = rows_to_dicts(conn.execute(f"SELECT id, name FROM options_{t} ORDER BY name").fetchall())
    options["pests"] = rows_to_dicts(conn.execute("""
        SELECT id, name, COALESCE(recommended_materials, '') AS recommended_materials
        FROM options_pests
        ORDER BY name
    """).fetchall())
    plans = rows_to_dicts(conn.execute("SELECT * FROM plans ORDER BY plan_date DESC, id DESC").fetchall())
    conn.close()

    payload = {
        "ok": True,
        "season": season_payload(season),
        "exported_at": current_timestamp(),
        "works": rows_to_dicts(work_rows),
        "money": money_rows,
        "materials": materials,
        "options": options,
        "plans": plans,
        "summary": {
            "works_count": len(work_rows),
            "money_count": len(money_rows),
            "money_total": sum(item["total"] for item in money_rows)
        }
    }
    return jsonify(payload)


# =========================
# IMPORT OLD DB
# =========================
@app.route("/api/import_old_db", methods=["POST"])
def import_old_db():
    uploaded = request.files.get("file")
    if not uploaded or not uploaded.filename:
        return jsonify({"ok": False, "error": "파일이 없습니다."}), 400

    temp_dir = tempfile.gettempdir()
    temp_path = os.path.join(temp_dir, f"old_worklog_{uuid.uuid4().hex}.db")

    try:
        uploaded.save(temp_path)
        result = import_old_db_into_current(temp_path)
        status = 200 if result.get("ok") else 400
        return jsonify(result), status
    finally:
        try:
            if os.path.exists(temp_path):
                os.remove(temp_path)
        except Exception:
            pass



# =========================
# INCOMES
# =========================
@app.route("/api/incomes", methods=["GET"])
def get_incomes():
    conn = db()
    season_id = (request.args.get("season_id") or "").strip()
    season = get_season_row(conn, season_id)

    sql = "SELECT * FROM incomes"
    params = ()
    if season:
        sql += " WHERE income_date BETWEEN ? AND ?"
        params = (season["start_date"], season["end_date"])

    rows = conn.execute(sql + " ORDER BY income_date DESC, id DESC", params).fetchall()
    conn.close()
    return jsonify(rows_to_dicts(rows))


@app.route("/api/incomes", methods=["POST"])
def create_income():
    data = request.get_json(force=True) or {}

    income_date = (data.get("income_date") or "").strip()
    income_type = (data.get("income_type") or "").strip()
    amount = safe_float(data.get("amount"), 0)
    method = (data.get("method") or "").strip()
    note = (data.get("note") or "").strip()

    if not income_date:
        return jsonify({"ok": False, "error": "수익 날짜가 없습니다."}), 400
    if not income_type:
        return jsonify({"ok": False, "error": "수익 구분이 없습니다."}), 400
    if amount <= 0:
        return jsonify({"ok": False, "error": "수익 금액이 올바르지 않습니다."}), 400

    conn = db()
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO incomes (income_date, income_type, amount, method, note, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (
            income_date,
            income_type,
            amount,
            method,
            note,
            current_timestamp()
        ))
        conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"ok": False, "error": f"수익 저장 오류: {e}"}), 500
    finally:
        conn.close()


@app.route("/api/incomes/<int:income_id>", methods=["DELETE"])
def delete_income(income_id):
    conn = db()
    try:
        conn.execute("DELETE FROM incomes WHERE id = ?", (income_id,))
        conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"ok": False, "error": f"수익 삭제 오류: {e}"}), 500
    finally:
        conn.close()


@app.route("/api/incomes/<int:income_id>", methods=["PUT"])
def update_income(income_id):
    data = request.get_json(force=True) or {}

    income_date = (data.get("income_date") or "").strip()
    income_type = (data.get("income_type") or "").strip()
    amount = safe_float(data.get("amount"), 0)
    method = (data.get("method") or "").strip()
    note = (data.get("note") or "").strip()

    if not income_date:
        return jsonify({"ok": False, "error": "수익 날짜가 없습니다."}), 400
    if not income_type:
        return jsonify({"ok": False, "error": "수익 구분이 없습니다."}), 400
    if amount <= 0:
        return jsonify({"ok": False, "error": "수익 금액이 올바르지 않습니다."}), 400

    conn = db()
    try:
        cur = conn.cursor()
        cur.execute("""
            UPDATE incomes
            SET income_date = ?,
                income_type = ?,
                amount = ?,
                method = ?,
                note = ?
            WHERE id = ?
        """, (
            income_date,
            income_type,
            amount,
            method,
            note,
            income_id
        ))
        conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"ok": False, "error": f"수익 수정 오류: {e}"}), 500
    finally:
        conn.close()


# =========================
# MONEY
# =========================
@app.route("/api/money", methods=["GET"])
def get_money():
    conn = db()
    season_id = (request.args.get("season_id") or "").strip()
    season = get_season_row(conn, season_id)
    sql, params = apply_season_filter_sql("SELECT * FROM works", season, "start_date")
    rows = conn.execute(sql + " ORDER BY start_date DESC, id DESC", params).fetchall()
    conn.close()

    result = []
    for row in rows:
        item = dict(row)
        memo = {}
        try:
            memo = json.loads(item.get("memo") or "{}")
        except Exception:
            memo = {}

        money = memo.get("money") or {}
        if not money:
            continue

        total_amount = float(money.get("total_amount") or money.get("amount") or 0)
        labor_total = float(money.get("labor_total") or 0)
        material_total = float(money.get("material_total") or 0)
        other_total = float(money.get("other_total") or 0)
        method = (money.get("method") or "").strip()
        installment_months = int(safe_float(money.get("installment_months"), 0))

        cash_amount = total_amount if method == "현금" else 0
        transfer_amount = total_amount if method == "계좌이체" else 0
        card_lump_amount = total_amount if method == "카드일시불" else 0
        card_install_amount = total_amount if method == "카드할부" else 0
        credit_amount = total_amount if method == "외상" else 0

        method_display = method
        if method == "카드할부" and installment_months > 0:
            method_display = f"카드할부 ({installment_months}개월)"

        result.append({
            "date": item.get("start_date", ""),
            "task_name": item.get("task_name", ""),
            "type": money.get("type", ""),
            "total": total_amount,
            "total_amount": total_amount,
            "labor_total": labor_total,
            "material_total": material_total,
            "other_total": other_total,
            "method": method,
            "method_display": method_display,
            "installment_months": installment_months,
            "cash_amount": cash_amount,
            "transfer_amount": transfer_amount,
            "card_lump_amount": card_lump_amount,
            "card_install_amount": card_install_amount,
            "credit_amount": credit_amount,
            "note": money.get("note", "")
        })

    return jsonify(result)

# =========================
# EXCEL EXPORT
# =========================
def build_excel_export_data(conn, season_id=''):
    season = get_season_row(conn, season_id)

    works_sql, works_params = apply_season_filter_sql("SELECT * FROM works", season, "start_date")
    works = rows_to_dicts(conn.execute(works_sql + " ORDER BY start_date DESC, id DESC", works_params).fetchall())

    income_sql = "SELECT * FROM incomes"
    income_params = ()
    if season:
        income_sql += " WHERE income_date BETWEEN ? AND ?"
        income_params = (season["start_date"], season["end_date"])
    incomes = rows_to_dicts(conn.execute(income_sql + " ORDER BY income_date DESC, id DESC", income_params).fetchall())

    work_rows = []
    money_rows = []

    for item in works:
        memo = parse_memo_json(item.get("memo") or "")
        materials = memo.get("materials") or []
        labor_rows = memo.get("labor_rows") or []
        money = memo.get("money") or {}

        materials_text = ", ".join([
            f"{(m.get('name') or '').strip()} {safe_float(m.get('qty'), 0):g}{(m.get('unit') or '').strip()} ({(m.get('action') or m.get('behavior') or '사용').strip() or '사용'})".strip()
            for m in materials if (m.get('name') or '').strip()
        ])

        labor_text = ", ".join([
            f"{(r.get('type') or '').strip()} {int(safe_float(r.get('count'), 0))}명 x {safe_float(r.get('price'), 0):,.0f}"
            for r in labor_rows
            if (r.get('type') or '').strip() or safe_float(r.get('count'), 0) or safe_float(r.get('price'), 0)
        ])

        total_amount = safe_float(money.get("total_amount"), 0) or safe_float(money.get("amount"), 0)
        labor_total = safe_float(money.get("labor_total"), 0)
        material_total = safe_float(money.get("material_total"), 0)
        other_total = safe_float(money.get("other_total"), 0)

        work_rows.append({
            "날짜": item.get("start_date", ""),
            "종료일": item.get("end_date", ""),
            "작업분류": item.get("task_category", ""),
            "세부작업": item.get("task_name", ""),
            "작물": item.get("crops", ""),
            "병충해": item.get("pests", ""),
            "사용기계": item.get("machines", ""),
            "작업시간": safe_float(item.get("work_hours"), 0),
            "날씨": item.get("weather", ""),
            "사용자재": materials_text,
            "인건비내역": labor_text,
            "자재비": material_total,
            "인건비": labor_total,
            "기타비": other_total,
            "비용구분": money.get("type", ""),
            "비용합계": total_amount,
            "결제방식": money.get("method", ""),
            "비용비고": money.get("note", ""),
            "작업메모": memo.get("memo_text", "")
        })

        if total_amount:
            money_rows.append({
                "날짜": item.get("start_date", ""),
                "이름": item.get("task_name", ""),
                "구분": money.get("type", ""),
                "금액": total_amount,
                "방식": money.get("method", ""),
                "비고": money.get("note", ""),
                "행종류": "지출"
            })

    for income in incomes:
        money_rows.append({
            "날짜": income.get("income_date", ""),
            "이름": income.get("income_type", ""),
            "구분": "수익",
            "금액": safe_float(income.get("amount"), 0),
            "방식": income.get("method", ""),
            "비고": income.get("note", ""),
            "행종류": "수익"
        })

    monthly_map = {}
    for row in money_rows:
        date_value = row.get("날짜", "")
        if not date_value or len(date_value) < 7:
            continue
        month_key = date_value[:7]
        if month_key not in monthly_map:
            monthly_map[month_key] = {"수익": 0, "지출": 0}
        if row.get("행종류") == "수익":
            monthly_map[month_key]["수익"] += safe_float(row.get("금액"), 0)
        else:
            monthly_map[month_key]["지출"] += safe_float(row.get("금액"), 0)

    monthly_rows = []
    for month_key in sorted(monthly_map.keys(), reverse=True):
        income_total = monthly_map[month_key]["수익"]
        expense_total = monthly_map[month_key]["지출"]
        monthly_rows.append({
            "월": month_key,
            "수익": income_total,
            "지출": expense_total,
            "순이익": income_total - expense_total
        })

    seasons = rows_to_dicts(conn.execute("SELECT * FROM seasons ORDER BY start_date DESC, id DESC").fetchall())
    season_rows = []
    for s in seasons:
        start_date = s.get("start_date", "")
        end_date = s.get("end_date", "")
        income_total = sum(
            safe_float(r.get("금액"), 0)
            for r in money_rows
            if r.get("행종류") == "수익" and start_date <= (r.get("날짜") or "") <= end_date
        )
        expense_total = sum(
            safe_float(r.get("금액"), 0)
            for r in money_rows
            if r.get("행종류") != "수익" and start_date <= (r.get("날짜") or "") <= end_date
        )
        season_rows.append({
            "시즌명": s.get("season_name", ""),
            "시작일": start_date,
            "종료일": end_date,
            "수익": income_total,
            "지출": expense_total,
            "순이익": income_total - expense_total,
            "비고": s.get("note", "")
        })

    total_income = sum(safe_float(row.get("금액"), 0) for row in money_rows if row.get("행종류") == "수익")
    total_expense = sum(safe_float(row.get("금액"), 0) for row in money_rows if row.get("행종류") != "수익")

    summary_rows = [
        {"항목": "정산범위", "값": (season.get("season_name") if season else "전체"), "비고": ""},
        {"항목": "총 수익", "값": total_income, "비고": ""},
        {"항목": "총 지출", "값": total_expense, "비고": ""},
        {"항목": "순이익", "값": total_income - total_expense, "비고": ""},
        {"항목": "작업 건수", "값": len(work_rows), "비고": ""},
        {"항목": "금전 건수", "값": len(money_rows), "비고": ""}
    ]

    return {
        "summary": summary_rows,
        "works": work_rows,
        "money": sorted(money_rows, key=lambda x: (x.get("날짜", ""), x.get("행종류", "")), reverse=True),
        "monthly": monthly_rows,
        "season_summary": season_rows
    }


def style_header_row(ws, row_idx=1):
    fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    font = Font(bold=True, color="FFFFFF")
    thin = Side(border_style="thin", color="D9E2F2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[row_idx]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


def auto_fit_columns(ws, min_width=10, max_width=42):
    for col_cells in ws.columns:
        first = col_cells[0]
        if getattr(first, 'column_letter', None) is None:
            continue
        max_length = 0
        col_letter = first.column_letter
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, min_width), max_width)


def apply_currency_format(ws, col_indexes, start_row=2):
    for col_idx in col_indexes:
        for row in ws.iter_rows(min_row=start_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '₩#,##0;[Red]-₩#,##0'
                    cell.alignment = Alignment(horizontal='right', vertical='center')


def write_sheet_rows(ws, title, headers, rows):
    ws.title = title
    ws.append(headers)
    style_header_row(ws, 1)

    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{ws.cell(1, ws.max_column).column_letter}{ws.max_row}"
    auto_fit_columns(ws)


def append_total_row(ws, label_col=1, amount_cols=None):
    if amount_cols is None:
        amount_cols = []
    last_data_row = ws.max_row
    total_row = last_data_row + 1
    ws.cell(total_row, label_col).value = "합계"
    ws.cell(total_row, label_col).font = Font(bold=True)

    fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    thin = Side(border_style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(total_row, col_idx)
        cell.fill = fill
        cell.border = border
        if col_idx in amount_cols:
            col_letter = ws.cell(1, col_idx).column_letter
            cell.value = f"=SUM({col_letter}2:{col_letter}{last_data_row})"
            cell.number_format = '₩#,##0;[Red]-₩#,##0'
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.font = Font(bold=True)
        elif col_idx != label_col:
            cell.value = ""


def style_money_rows(ws):
    income_fill = PatternFill(fill_type='solid', fgColor='EAF7EA')
    expense_fill = PatternFill(fill_type='solid', fgColor='FDECEC')
    for row_idx in range(2, ws.max_row + 1):
        row_type = ws.cell(row_idx, 7).value
        fill = income_fill if row_type == '수익' else expense_fill
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row_idx, col_idx).fill = fill
        amount_cell = ws.cell(row_idx, 4)
        amount_cell.font = Font(bold=True, color='008000' if row_type == '수익' else 'C00000')


def set_summary_dashboard(ws, export_data):
    ws.title = '요약'
    ws.sheet_view.showGridLines = False
    for col, width in {'A':16,'B':16,'C':16,'D':16,'E':16,'F':16,'G':16,'H':16,'I':16,'J':16,'K':16,'L':16}.items():
        ws.column_dimensions[col].width = width

    ws.merge_cells('A1:L2')
    title = ws['A1']
    title.value = '작업일지 운영 보고서'
    title.fill = PatternFill(fill_type='solid', fgColor='1F4E78')
    title.font = Font(bold=True, color='FFFFFF', size=20)
    title.alignment = Alignment(horizontal='center', vertical='center')

    scope_name = '전체'
    for row in export_data['summary']:
        if row.get('항목') == '정산범위':
            scope_name = row.get('값') or '전체'
            break

    ws.merge_cells('A3:L3')
    meta = ws['A3']
    meta.value = f"정산범위: {scope_name}   |   생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    meta.font = Font(bold=True, color='1F1F1F', size=11)
    meta.alignment = Alignment(horizontal='center', vertical='center')

    summary_map = {row.get('항목'): row.get('값') for row in export_data['summary']}
    total_income = safe_float(summary_map.get('총 수익'), 0)
    total_expense = safe_float(summary_map.get('총 지출'), 0)
    net_profit = safe_float(summary_map.get('순이익'), 0)
    work_count = int(safe_float(summary_map.get('작업 건수'), 0))
    money_count = int(safe_float(summary_map.get('금전 건수'), 0))
    month_count = len(export_data.get('monthly') or [])
    season_count = len(export_data.get('season_summary') or [])
    status_text = '흑자' if net_profit > 0 else ('적자' if net_profit < 0 else '균형')

    cards = [
        ('A5:C8', '총 수익', total_income, 'income'),
        ('D5:F8', '총 지출', total_expense, 'expense'),
        ('G5:I8', '순이익', net_profit, 'net'),
        ('J5:L6', '작업 건수', work_count, 'count'),
        ('J7:L8', '금전 건수', money_count, 'count'),
        ('A9:D10', '월 수', month_count, 'sub'),
        ('E9:H10', '시즌 수', season_count, 'sub'),
        ('I9:L10', '현재 상태', status_text, 'status')
    ]
    fills = {
        'income': 'E2F0D9',
        'expense': 'FCE4D6',
        'net': 'FFF2CC',
        'count': 'D9E2F3',
        'sub': 'F3F3F3',
        'status': 'EAD1DC' if net_profit < 0 else 'D9EAD3'
    }
    font_colors = {
        'income': '006100',
        'expense': 'C00000',
        'net': 'C00000' if net_profit < 0 else '1F4E78',
        'count': '1F1F1F',
        'sub': '7F6000',
        'status': 'C00000' if net_profit < 0 else '006100'
    }
    thin = Side(border_style='thin', color='D9D9D9')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for rng, label, value, kind in cards:
        ws.merge_cells(rng)
        cell = ws[rng.split(':')[0]]
        cell.value = f"{label}\n{value if isinstance(value, str) else ""}"
        if not isinstance(value, str):
            if kind in ('income', 'expense', 'net'):
                cell.value = f"{label}\n₩ {value:,.0f}"
            else:
                cell.value = f"{label}\n{int(value):,}"
        cell.fill = PatternFill(fill_type='solid', fgColor=fills[kind])
        cell.font = Font(bold=True, color=font_colors[kind], size=16)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    ws['A12'] = '월 수'
    ws['B12'] = month_count
    ws['C12'] = '월별정산 시트 참조'
    ws['A13'] = '시즌 수'
    ws['B13'] = season_count
    ws['C13'] = '시즌정산 시트 참조'
    ws['A14'] = '작업 건수'
    ws['B14'] = work_count
    ws['C14'] = '작업일지 시트 참조'
    for row in ws.iter_rows(min_row=12, max_row=14, min_col=1, max_col=3):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if cell.column == 1:
                cell.fill = PatternFill(fill_type='solid', fgColor='EAF2FF')
                cell.font = Font(bold=True)


def add_summary_charts(ws_summary, ws_monthly, ws_season):
    monthly_data_end = ws_monthly.max_row - 1 if ws_monthly.max_row >= 3 else ws_monthly.max_row
    if monthly_data_end >= 2:
        bar = BarChart()
        bar.type = 'col'
        bar.style = 10
        bar.title = '월별 수익 / 지출'
        bar.y_axis.title = '금액'
        bar.x_axis.title = '월'
        data = Reference(ws_monthly, min_col=2, max_col=3, min_row=1, max_row=monthly_data_end)
        cats = Reference(ws_monthly, min_col=1, min_row=2, max_row=monthly_data_end)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        bar.height = 10
        bar.width = 16
        ws_summary.add_chart(bar, 'A18')

        line = LineChart()
        line.style = 2
        line.title = '월별 순이익'
        line.y_axis.title = '순이익'
        line.x_axis.title = '월'
        data2 = Reference(ws_monthly, min_col=4, max_col=4, min_row=1, max_row=monthly_data_end)
        cats2 = Reference(ws_monthly, min_col=1, min_row=2, max_row=monthly_data_end)
        line.add_data(data2, titles_from_data=True)
        line.set_categories(cats2)
        line.height = 10
        line.width = 16
        ws_summary.add_chart(line, 'G18')

    season_data_end = ws_season.max_row - 1 if ws_season.max_row >= 3 else ws_season.max_row
    if season_data_end >= 2:
        season_chart = BarChart()
        season_chart.type = 'col'
        season_chart.style = 12
        season_chart.title = '시즌별 순이익'
        season_chart.y_axis.title = '순이익'
        season_chart.x_axis.title = '시즌'
        data3 = Reference(ws_season, min_col=6, max_col=6, min_row=1, max_row=season_data_end)
        cats3 = Reference(ws_season, min_col=1, min_row=2, max_row=season_data_end)
        season_chart.add_data(data3, titles_from_data=True)
        season_chart.set_categories(cats3)
        season_chart.height = 10
        season_chart.width = 16
        ws_summary.add_chart(season_chart, 'A36')


def build_excel_file(export_data):
    wb = Workbook()

    ws0 = wb.active
    set_summary_dashboard(ws0, export_data)

    ws1 = wb.create_sheet('작업일지')
    write_sheet_rows(
        ws1,
        '작업일지',
        ['날짜', '종료일', '작업분류', '세부작업', '작물', '병충해', '사용기계', '작업시간', '날씨', '사용자재', '인건비내역', '자재비', '인건비', '기타비', '비용구분', '비용합계', '결제방식', '비용비고', '작업메모'],
        export_data['works']
    )
    apply_currency_format(ws1, [12, 13, 14, 16], start_row=2)
    append_total_row(ws1, label_col=1, amount_cols=[12, 13, 14, 16])

    ws2 = wb.create_sheet('금전내역')
    write_sheet_rows(
        ws2,
        '금전내역',
        ['날짜', '이름', '구분', '금액', '방식', '비고', '행종류'],
        export_data['money']
    )
    apply_currency_format(ws2, [4], start_row=2)
    style_money_rows(ws2)
    append_total_row(ws2, label_col=1, amount_cols=[4])

    ws3 = wb.create_sheet('월별정산')
    write_sheet_rows(
        ws3,
        '월별정산',
        ['월', '수익', '지출', '순이익'],
        export_data['monthly']
    )
    apply_currency_format(ws3, [2, 3, 4], start_row=2)
    append_total_row(ws3, label_col=1, amount_cols=[2, 3, 4])

    ws4 = wb.create_sheet('시즌정산')
    write_sheet_rows(
        ws4,
        '시즌정산',
        ['시즌명', '시작일', '종료일', '수익', '지출', '순이익', '비고'],
        export_data['season_summary']
    )
    apply_currency_format(ws4, [4, 5, 6], start_row=2)
    append_total_row(ws4, label_col=1, amount_cols=[4, 5, 6])

    ws3.column_dimensions['D'].width = 16
    ws4.column_dimensions['F'].width = 16
    ws4.column_dimensions['G'].width = 20

    add_summary_charts(ws0, ws3, ws4)

    return wb


@app.route("/api/export_excel", methods=["GET"])
def export_excel():
    season_id = (request.args.get("season_id") or "").strip()

    conn = db()
    try:
        export_data = build_excel_export_data(conn, season_id)
    finally:
        conn.close()

    wb = build_excel_file(export_data)
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    scope_name = season_id or "all"
    filename = f"worklog_report_{scope_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if __name__ == "__main__":
    app.run(debug=True)

# =========================
# FAVORITE WORK TEMPLATES
# =========================
@app.route("/api/favorite-work-templates", methods=["GET"])
def get_favorite_work_templates():
    conn = db()
    rows = conn.execute("""
        SELECT id, name, template_json, created_at, updated_at
        FROM favorite_work_templates
        ORDER BY name COLLATE NOCASE ASC, id DESC
    """).fetchall()
    conn.close()

    result = []
    for row in rows_to_dicts(rows):
        template = {}
        try:
            template = json.loads(row.get("template_json") or "{}")
            if not isinstance(template, dict):
                template = {}
        except Exception:
            template = {}
        result.append({
            "id": row.get("id"),
            "name": row.get("name") or "",
            "template": template,
            "created_at": row.get("created_at") or "",
            "updated_at": row.get("updated_at") or ""
        })
    return jsonify(result)


@app.route("/api/favorite-work-templates", methods=["POST"])
def save_favorite_work_template():
    data = request.get_json(force=True) or {}
    name = normalize_name(data.get("name"))
    template = data.get("template") or {}

    if not name:
        return jsonify({"ok": False, "error": "즐겨찾기 이름이 없습니다."}), 400
    if not isinstance(template, dict):
        return jsonify({"ok": False, "error": "즐겨찾기 데이터 형식이 올바르지 않습니다."}), 400

    now = current_timestamp()
    template_json = json.dumps(template, ensure_ascii=False)

    conn = db()
    cur = conn.cursor()
    existing = cur.execute("SELECT id FROM favorite_work_templates WHERE name = ?", (name,)).fetchone()

    if existing:
        cur.execute(
            "UPDATE favorite_work_templates SET template_json = ?, updated_at = ? WHERE id = ?",
            (template_json, now, existing["id"])
        )
        favorite_id = existing["id"]
    else:
        cur.execute(
            "INSERT INTO favorite_work_templates (name, template_json, created_at, updated_at) VALUES (?, ?, ?, ?)",
            (name, template_json, now, now)
        )
        favorite_id = cur.lastrowid

    conn.commit()
    conn.close()
    return jsonify({"ok": True, "id": favorite_id, "name": name})


@app.route("/api/favorite-work-templates/<int:favorite_id>", methods=["DELETE"])
def delete_favorite_work_template(favorite_id):
    conn = db()
    cur = conn.cursor()
    cur.execute("DELETE FROM favorite_work_templates WHERE id = ?", (favorite_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})
